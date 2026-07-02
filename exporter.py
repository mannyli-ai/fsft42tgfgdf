"""把翻譯結果匯出成 xlsx。連結欄是真正的超連結，Excel 打開直接可點。"""

from __future__ import annotations

import io
from urllib.parse import urlparse

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

HEADERS = ["日期", "原始標題", "翻譯標題", "原文連結", "Techmeme 連結"]
COLUMN_WIDTHS = [12, 60, 60, 28, 16]

FONT_NAME = "Arial"
HYPERLINK_FONT = Font(name=FONT_NAME, color="0563C1", underline="single")
HEADER_FONT = Font(name=FONT_NAME, bold=True)
BODY_FONT = Font(name=FONT_NAME)


def _domain(url: str) -> str:
    """原文連結的顯示文字用網域，掃表格時一眼看得出來源。"""
    try:
        host = urlparse(url).netloc
        return host.removeprefix("www.") or url
    except Exception:
        return url


def build_xlsx(items: list[dict]) -> bytes:
    """items 需要的欄位：date, title, title_zh, article_url, techmeme_url。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Techmeme"

    ws.append(HEADERS)
    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center")

    for it in items:
        row_idx = ws.max_row + 1
        ws.cell(row=row_idx, column=1, value=it.get("date", "")).font = BODY_FONT
        ws.cell(row=row_idx, column=2, value=it.get("title", "")).font = BODY_FONT
        ws.cell(row=row_idx, column=3, value=it.get("title_zh", "")).font = BODY_FONT

        article_url = it.get("article_url", "")
        c4 = ws.cell(row=row_idx, column=4, value=_domain(article_url))
        if article_url:
            c4.hyperlink = article_url
            c4.font = HYPERLINK_FONT

        tm_url = it.get("techmeme_url", "")
        c5 = ws.cell(row=row_idx, column=5, value="Techmeme")
        if tm_url:
            c5.hyperlink = tm_url
            c5.font = HYPERLINK_FONT

    for i, width in enumerate(COLUMN_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
